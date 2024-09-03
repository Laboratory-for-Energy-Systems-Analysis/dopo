
import brightway2 as bw
import bw2data as bd
import bw2analyzer as ba

#reduce?
import ast
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import dopo
from dopo import*

import pandas as pd
from openpyxl import load_workbook

import pandas as pd

import re
import pandas as pd

from dopo import small_inputs_to_other_column
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart import BarChart, Reference

def _lca_scores_compare(database_dict, method_dict):
    # Dictionary to store DataFrames for each sector
    sector_dataframes = {}

    # Labels for the DataFrame columns
    labels = [
        "activity",
        "activity key",
        "reference product",
        "location",
        "method",
        "method unit",
        "total",
    ]

    # Loop through each sector in the database_dict
    for sector, sector_data in database_dict.items():
        # Initialize a dictionary to hold DataFrames for each method in the current sector
        method_dataframes = {}

        # Loop through each method in method_dict
        for meth_key, meth_info in method_dict.items():
            data = []  # Initialize a new list to hold data for the current method
            
            # Extract the 'method name' tuple from the current method info
            method_name = meth_info['method name']
            method_short_name=meth_info['short name']
            method_unit = meth_info['unit']

            # Now loop through each activity in the sector
            for act in sector_data['activities']:
                # Ensure the activity is an instance of the expected class
                if not isinstance(act, bd.backends.peewee.proxies.Activity):
                    raise ValueError("`activities` must be an iterable of `Activity` instances")
                
                # Perform LCA calculations
                lca = bw.LCA({act: 1}, method_name)
                lca.lci()
                lca.lcia()
                
                # Collect data for the current activity and method
                data.append([
                    act["name"],
                    act.key,
                    act.get("reference product"),
                    act.get("location", "")[:25],
                    method_short_name,
                    method_unit,
                    lca.score,
                ])
            
            # Convert the data list to a DataFrame and store it in the sector's dictionary
            method_dataframes[method_short_name] = pd.DataFrame(data, columns=labels)

        # Store the method_dataframes dictionary in the sector_dataframes dictionary
        sector_dataframes[sector] = method_dataframes

    # Now `sector_dataframes` is a dictionary where each key is a sector, and the value is another dictionary with method names and their corresponding DataFrames
    return sector_dataframes


import pandas as pd

def _relative_changes_df(database_dict_eco, database_dict_premise, method_dict):

    ecoinvent_scores = _lca_scores_compare(database_dict_eco, method_dict)
    premise_scores = _lca_scores_compare(database_dict_premise, method_dict)

    relative_dict = {}

    # Iterate over sectors
    for sector_key in ecoinvent_scores:
        # Initialize the sector key in the output dictionary
        if sector_key not in relative_dict:
            relative_dict[sector_key] = {}

        # Iterate over methods within the sector
        for method_key in ecoinvent_scores[sector_key]:
            # Check if the method_key exists in both dictionaries to avoid KeyError
            if method_key in premise_scores.get(sector_key, {}):
                # Get the corresponding DataFrames
                df_ei = ecoinvent_scores[sector_key][method_key]
                df_premise = premise_scores[sector_key][method_key]

                #print(df_ei['activity key'])
                #print(df_premise)

                # Split the 'activity key' to extract the second part
                df_ei['activity_code'] = df_ei['activity key'].apply(lambda x: x[1])  # Access the second element of the tuple
                df_premise['activity_code'] = df_premise['activity key'].apply(lambda x: x[1])

                # Merge the two dataframes based on the activity code and method name
                merged_df = pd.merge(df_ei, df_premise, on=['activity_code', 'method'], suffixes=('_ei', '_premise'))

                # Calculate the relative change
                merged_df['relative_change'] = ((merged_df['total_premise'] - merged_df['total_ei']) / merged_df['total_ei']) * 100

                # Store the result in the dictionary
                relative_dict[sector_key][method_key] = merged_df

    return relative_dict

def relative_changes_db(database_dict_eco, database_dict_premise, method_dict, excel_file):
    relative_dict = _relative_changes_df(database_dict_eco, database_dict_premise, method_dict)
    
    # Load existing workbook and get existing sheet names
    try:
        book = load_workbook(excel_file)
        existing_sheets = book.sheetnames
    except FileNotFoundError:
        # If the file does not exist, we will create a new one, so no need to check existing sheets
        existing_sheets = []
    
    column_positions = {}  # stores the indexes of columns for plotting
    
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        for sector in relative_dict.keys():
            relative_changes = relative_dict[sector]
            
            for method, table in relative_changes.items():
                # Create a DataFrame for the current LCA score table
                df = pd.DataFrame(table)

                # Add sector marker
                df = _add_sector_marker(df, sector) #!! ADJUST      

                # Sort the DataFrame by 'relative_change' from largest negative to largest positive
                df = df.sort_values(by='relative_change', ascending=False)

                # Add a 'rank' column based on the 'relative_change', ranking from most negative to least negative
                df['rank'] = df['relative_change'].rank(ascending=False, method='dense').astype(int)
    
                # Get the index values of columns
                columns_of_interest = ["rank", "relative_change", "method", "method unit_ei"]
                positions = {col: df.columns.get_loc(col) for col in columns_of_interest if col in df.columns}
                column_positions[f"{sector}_comparison_{method}"] = positions

                # Generate worksheet name
                worksheet_name = f"{sector}_comparison_{method}"
                if len(worksheet_name) > 31:
                    worksheet_name = worksheet_name[:31]
                
                # Ensure unique sheet name
                original_worksheet_name = worksheet_name
                counter = 1
                while worksheet_name in existing_sheets:
                    worksheet_name = f"{original_worksheet_name}_{counter}"
                    if len(worksheet_name) > 31:
                        worksheet_name = worksheet_name[:31]
                    counter += 1

                # Save the DataFrame to the Excel file in a new worksheet
                df.to_excel(writer, sheet_name=worksheet_name, index=False)
    
    return column_positions


def _categorize_sheets_by_sector_comparison(file_path):
    # Load the workbook
    workbook = load_workbook(filename=file_path, read_only=True)
    
    # Initialize a dictionary to hold sectors and their corresponding sheet names
    worksheet_dict = {}
    
    # Iterate over all sheet names in the workbook
    for sheet_name in workbook.sheetnames:
        # Skip combined sector sheets (assuming these sheets don't have an underscore)
        if '_comparison' not in sheet_name:
            continue
        
        # Split the sheet name to extract the sector (assumes sector is the first part)
        sector = sheet_name.split('_')[0]
        
        # Add the sheet name to the corresponding sector in the dictionary
        if sector in worksheet_dict:
            worksheet_dict[sector].append(sheet_name)
        else:
            worksheet_dict[sector] = [sheet_name]
    
    return worksheet_dict


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines

def barchart_compare_db_xcl(filename, index_positions, current_row_stacked_bar): #, index_positions=None):
      
    worksheet_dict =_categorize_sheets_by_sector_comparison(file_path=filename)
    # Load the workbook and select the sheet
    wb = load_workbook(filename)

    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)  
        
        # Initial position for the first chart
        current_col = 1  # Start placing charts from column 1
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        charts_per_row = 3  # Number of charts per row
        current_row = current_row_stacked_bar + chart_height
    
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in index_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = index_positions[matching_key]

            # Find min_row, max_row and max_column
            min_col_data = positions.get("relative_change", None) + 1 #15
            rank_col = positions.get("rank", None) + 1 #17
            method_col = positions.get("method", None) + 1 #5
            method_unit_col = positions.get("method unit_ei", None) + 1 #6

            # Create a bar chart
            chart = BarChart()
            chart.type="bar"
            chart.style=2
            chart.overlap= 100

            # Set the data for the chart
            data = Reference(ws, min_col=min_col_data, min_row=2, max_row=ws.max_row)
            categories = Reference(ws, min_col=rank_col, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)

            # Modify each series in the chart to disable the inversion of negative values 
            for series in chart.series:
                series.invertIfNegative = False

            # Y-axis (categories) settings
            chart.y_axis.majorGridlines = ChartLines()
            chart.y_axis.delete = False
            chart.y_axis.tickLblPos = "low"

            # X-axis (values) settings
            chart.x_axis.majorGridlines = None

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.title = f"{sector} - Database relatvie comparison for method: {method_value}"


            chart.x_axis.title = "Activity"
            chart.y_axis.title = "Relative Change (%)"

            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 
            chart.legend.overlay = False

            # Adjust chart dimensions
            chart.width = 30  # Increased width for better readability
            chart.height = 15

            # Adjust chart dimensions
            chart.width = 20
            chart.height = 14
            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)

            # Update position for the next chart
            current_col += chart_width +1 
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)

            # Add the chart to a new worksheet
            # new_sheet = wb.create_sheet(title="LCA Chart")
            # new_sheet.add_chart(chart, "A1")

    # Save the workbook
    wb.save(filename)

    print(f"Results and chart saved to {filename}")
