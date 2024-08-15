# Functions for dopo in excel
# dependencies

import re
import pandas as pd

import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart import BarChart, Reference

def process_yaml_files(files_dict, database):
    '''
    - Runs through the files_dict reading the defined filters in the yaml files.
    - With another function a list that contains the filtered activities is created from the chosen database.
    - This activity list is saved within the corresponding key (sector) in the dictionary main_dict which is based on the files_dict.

    :param files_dict: dictionary of dictionaries. It should hold the yaml file path and the title in the first row of the yaml file. 
                        Like so: files_dict['Cement']={'yaml': 'yamls\cement_small.yaml', 'yaml identifier': 'Cement'}
    :param database: premise or ecoinvent database of choice.

    It returns an updated dictionary which contains filtered activity lists for each sector.
    '''

    main_dict=files_dict.copy()
    for key, value in main_dict.items():
        yaml_file = value['yaml']
        yaml_identifier = value['yaml identifier']
        
        # Generate the sector activities
        sector_activities = generate_sets_from_filters(yaml_file, database=database)
        
        # Convert the set of activities to a list
        activities_list = list(sector_activities[yaml_identifier])
        
        # Add to the sectors_dict
        main_dict[key]['activities'] = activities_list
        
    return main_dict

def sector_lca_scores(main_dict, method_dict):
    '''
    Generates the LCA score tables for activity list of each sector.
    The tables contain total scores and cpc input contributions.
    This is done by each method defined in the method dictionary.

    :param main_dict: dictionary which is returned by process_yaml_files function
    :param method_dict: dictionary which is created with MethodFinder class

    It returns the main dictionary updated as scores dictionary which also holds the former information for each sector.
    The LCA scores are stored by method name in the respective sector dictionary within the main dictionary.
    '''

    # Initialize scores_dict as a copy of main_dict
    scores_dict = main_dict.copy()

    # Loop through each sector in main_dict
    for sector in scores_dict.keys():
        # Extract activities for the current sector
        sector_activities = scores_dict[sector]['activities']
        
        # Calculate LCA scores using the specified method
        lca_scores = compare_activities_multiple_methods(
            activities_list=sector_activities,
            methods=method_dict,
            identifier=sector,
            mode='absolute'
        )
        
        # Apply the small_inputs_to_other_column function with the cutoff value
        lca_scores = small_inputs_to_other_column(lca_scores, cutoff=0.02)
        
        # Save the LCA scores to the scores_dict
        scores_dict[sector]['lca_scores'] = lca_scores

    return scores_dict

# -----------------------------------------
# CREATING EXCEL SHEETS WITH LCA TABLES
# -----------------------------------------

def sector_lca_scores_to_excel_and_column_positions(scores_dict, excel_file_name):
    """ 
    What it does:
        - Creates a dataframe for each method and sector from the lca scores dictionary
        - Before storing each df in a worksheet in an excel file it:
                - shortens the column labels of the input (removing cpc code)
                - adds a sector name marker for keeping track in excel (when plotting can use it for labeling)
                - adds statistics for plotting
                - creates a dictionary which holds the indexes to the columns we need to call for plotting, this makes it dynamic. Otherwise need to hardcode index column number for openpxyl.
    What it returns:
        - Returns the index positions dictionary where the key is "sector_method"
        - Creates excel file as defined by user
    """

    # Prepare to save each LCA score table to a different worksheet in the same Excel file
    excel_file = excel_file_name
    column_positions = {} #stores the indexes of columns for plotting
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for sector in scores_dict.keys():
            lca_scores = scores_dict[sector]['lca_scores']
            for method, table in lca_scores.items():
                # Create a DataFrame for the current LCA score table
                df = pd.DataFrame(table)

                # Add sector marker
                df = add_sector_marker(df, sector) #!! ADJUST POSITION

                # Add statistics to the DataFrame
                df = add_statistics(df)

                # Get the index values of columns
                columns_of_interest = ["total", "rank", "mean", "2std_abv", "2std_blw", "q1", "q3", "method", "method unit"]
                positions = {col: df.columns.get_loc(col) for col in columns_of_interest if col in df.columns}
                column_positions[method] = positions

                # Find the first input column and add it to the positions dictionary
                first_input_col_index = find_first_input_column(df)
                if first_input_col_index is not None:
                    positions["first_input"] = first_input_col_index
                
                # Store the positions for this method
                column_positions[method] = positions

                # remove cpc from input labels
                df = clean_column_labels(df)

                # Generate a worksheet name
                worksheet_name = f"{method}" #f"{sector}_{method}"
                if len(worksheet_name) > 31:
                    worksheet_name = worksheet_name[:31]
                    
                # Save the DataFrame to the Excel file in a new worksheet
                df.to_excel(writer, sheet_name=worksheet_name, index=False)
        return column_positions
    

def add_statistics(df, column_name='total'):

    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions

    It adds statistical indicators to a dataframe based on total column which are used for plotting.

    returns updated dataframe
    '''

    #Need a rank row to plot the total LCA scores in descending order (satter opepyxl function takes in non categorial values)
    df['rank'] = df[column_name].rank(method="first", ascending="False")

    # Calculate mean, standard deviation, and IQR
    df['mean'] = df[column_name].mean()
    df['2std_abv'] = df['mean'] + df[column_name].std() * 2
    df['2std_blw'] = df['mean'] - df[column_name].std() * 2
    df['q1'] = df[column_name].quantile(0.25)
    df['q3'] = df[column_name].quantile(0.75)
    
    # Reorder the columns to place the new columns after 'total'
    cols = df.columns.tolist()
    total_index = cols.index(column_name) + 1
    new_cols = ['rank', 'mean', '2std_abv', '2std_blw', 'q1', 'q3']
    cols = cols[:total_index] + new_cols + cols[total_index:-len(new_cols)]
    
    return df[cols]


def find_first_input_column(df):
    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions. Needs to be called before clean_column_labels function.
    Detects the first column in the dataframe which contains input contribution data and saves its index. 
    This is relevant for calling the right column for defining the to be plotted data dynamically as not all dataframes have the same column order (some contain "direct emissions" for instance).
    '''
    
    def clean_label(label):
        return label if label is not None else 'Unnamed'
    
    # Apply the cleaning function to all column names
    df.columns = [clean_label(col) for col in df.columns]
    
    # Regular expression pattern to match "Number: Name" or "Number Name"
    pattern = r'^\d+:\s*'
    
    for idx, column in enumerate(df.columns):
        if (column is not None and re.match(pattern, column)) or column == 'Unnamed' or column == 'direct emissions':
            return idx

    return None

def clean_column_labels(df):

    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions. Needs to be run after find_first_input_column.

    It removes unnecessary numbers in the column header.

    Returns df with formated column labels.
    '''
    # Function to remove numbers and colon from column names
    def clean_label(label):
        if label is None:
            return 'Unnamed'  # or return 'Unnamed' if you prefer a placeholder
        return re.sub(r'^\d+:\s*', '', str(label))

    # Apply the cleaning function to all column names
    df.columns = [clean_label(col) for col in df.columns]

    return df

def add_sector_marker(df, sector):
'''
    It is called in the function sector_lca_scores_to_excel_and_column_positions.

    It adds information about the sector for titel and labeling in plotting.

    Returns df with added column.
    '''
    
    # Add sector marker column
    df['sector']=str(sector) # potentially remove!
    # Reorder the columns to move 'sector' after 'product'
    columns = list(df.columns)
    product_index = columns.index('product')
    # Insert 'sector' after 'product'
    columns.insert(product_index + 1, columns.pop(columns.index('sector')))
    # Reassign the DataFrame with the new column order
    df = df[columns]
    return df


# IN EXCEL
def categorize_sheets_by_sector(file_path):
    # Load the workbook
    workbook = load_workbook(filename=file_path, read_only=True)
    
    # Initialize a dictionary to hold sectors and their corresponding sheet names
    worksheet_dict = {}
    
    # Iterate over all sheet names in the workbook
    for sheet_name in workbook.sheetnames:
        # Split the sheet name to extract the sector (assumes sector is the first part)
        sector = sheet_name.split('_')[0]
        
        # Add the sheet name to the corresponding sector in the dictionary
        if sector in worksheet_dict:
            worksheet_dict[sector].append(sheet_name)
        else:
            worksheet_dict[sector] = [sheet_name]
    
    return worksheet_dict


# ----
#PLOTS
# ----

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

def dot_plots(filepath_workbook, worksheet_dict, index_positions):
    
    # Load the workbook
    wb = load_workbook(filepath_workbook)
    
    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)        
                
        # Initial position for the first chart
        current_row = 1  # Start placing charts from row 1
        current_col = 1  # Start placing charts from column 1
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        charts_per_row = 3  # Number of charts per row
        
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find min_row, max_row and max_column
            max_row = ws.max_row
            max_column = ws.max_column
            min_row = 1

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in index_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = index_positions[matching_key]
            total_col = positions.get("total", None) + 1
            rank_col = positions.get("rank", None) + 1
            mean_col = positions.get("mean", None) + 1
            std_adv_col = positions.get("2std_abv", None) + 1
            std_blw_col = positions.get("2std_blw", None) + 1
            q1_col = positions.get("q1", None) + 1
            q3_col = positions.get("q3", None) + 1
            method_col = positions.get("method", None) + 1
            method_unit_col = positions.get("method unit", None) + 1
            
            # Ensure that all required columns are present
            if None in [total_col, rank_col, mean_col, std_adv_col, std_blw_col, q1_col, q3_col, method_col, method_unit_col]:
                print(f"Warning: Missing columns in worksheet '{worksheet_name}' for sector '{sector}'. Skipping...")
                continue
            
            # Create a ScatterChart (or other chart type as needed)
            chart = ScatterChart()

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"{method_value} LCA scores for {sector} sector" 
            
            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.y_axis.title = f"{method_unit_value}"
            chart.x_axis.title = 'activity rank'
            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 

            # Define the data range for the chart
            y_values = Reference(ws, min_col=total_col, min_row=min_row, max_row=max_row)
            x_values = Reference(ws, min_col=rank_col, min_row=min_row, max_row=max_row)

            # Create a series and add it to the chart
            series = Series(y_values, x_values, title_from_data=True)
            chart.series.append(series)
            chart.style = 9

            # Customize the series to show only markers (dots)
            series.marker.symbol = "circle"
            series.marker.size = 5
            series.graphicalProperties.line.noFill = True

            # ADJUST X-AXIS
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.majorGridlines = None 
            chart.x_axis.tickMarkSkip = 1  # Show all tick marks, this adresses the tick lines 
            chart.x_axis.tickLblSkip = 1  # Show all labels, doesnt work

            chart.x_axis.scaling.orientation = "minMax"
            chart.x_axis.crosses = "autoZero"
            chart.x_axis.axPos = "b"
            chart.x_axis.delete = False

            # ADJUST Y-AXIS
            chart.y_axis.tickLblPos = "nextTo"  # Position the labels next to the tick marks
            chart.y_axis.delete = False  # Ensure axis is not deleted
            chart.y_axis.number_format = '0.00000'
            chart.y_axis.majorGridlines = None 

            # ADD STATS
            # MEAN
            mean_y = Reference(ws, min_col=mean_col, min_row=min_row, max_row=max_row)
            mean_series = Series(mean_y, x_values, title_from_data="True")
            chart.series.append(mean_series)
            mean_series.marker.symbol = "none"  # No markers, just a line
            mean_series.graphicalProperties.line.solidFill = "FF0000"  # Red line for mean value
            mean_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)

            # IQR
            iqr1 = Reference(ws, min_col=q1_col, min_row=min_row, max_row=max_row)
            iqr3 = Reference(ws, min_col=q3_col, min_row=min_row, max_row=max_row)
            iqr1_series = Series(iqr1, x_values, title_from_data="True")
            iqr3_series = Series(iqr3, x_values, title_from_data="True")
            chart.series.append(iqr1_series)
            chart.series.append(iqr3_series)
            iqr1_series.marker.symbol = "none"  # No markers, just a line
            iqr3_series.marker.symbol = "none"
            iqr1_series.graphicalProperties.line.solidFill = "6082B6"  # Blue line 
            iqr3_series.graphicalProperties.line.solidFill = "6082B6"  
            iqr1_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)
            iqr3_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)

            # STD
            std_abv = Reference(ws, min_col=std_adv_col, min_row=min_row, max_row=max_row)
            std_blw = Reference(ws, min_col=std_blw_col, min_row=min_row, max_row=max_row)
            std_abv_series = Series(std_abv, x_values, title_from_data="True")
            std_blw_series = Series(std_blw, x_values, title_from_data="True")
            chart.series.append(std_abv_series)
            chart.series.append(std_blw_series)
            std_abv_series.marker.symbol = "none"  # No markers, just a line
            std_blw_series.marker.symbol = "none"
            std_abv_series.graphicalProperties.line.solidFill = "FFC300"  # yellow line
            std_blw_series.graphicalProperties.line.solidFill = "FFC300"  
            std_abv_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)
            std_blw_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)

            # Set legend position to the right of the plot area
            chart.legend.position = 'r'  # 'r' for right
            chart.legend.overlay = False

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)
            
            # Update position for the next chart
            current_col += chart_width +1 
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)

    wb.save(filepath_workbook)
    return current_row


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def stacked_bars(filepath_workbook, worksheet_dict, current_row_dot_plot):
    # Load the workbook
    wb = load_workbook(filepath_workbook)
    
    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)
                
        # Initial position for the first chart
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        current_row = current_row_dot_plot + chart_height # Start placing charts from row where dot plots have left of
        current_col = 1  # Start placing charts from column 1
        charts_per_row = 3  # Number of charts per row
        
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in index_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = index_positions[matching_key]

            # Find min_row, max_row and max_column
            max_row = ws.max_row
            max_column = ws.max_column
            input_min_col = positions.get("first_input", None) + 1
            rank_col = positions.get("rank", None) + 1
            method_col = positions.get("method", None) + 1
            method_unit_col = positions.get("method unit", None) + 1

            chart = BarChart()
            chart.type = "bar"
            chart.style = 2
            chart.grouping = "stacked"
            chart.overlap = 100

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"{sector} sector inputs contributions to {method_value}"

            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.y_axis.title = f"{method_unit_value}"
            
            chart.x_axis.title = 'activity index'

            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 
            chart.legend.overlay = False

            # Define data
            data = Reference(ws, min_col=input_min_col, min_row=1, max_row=max_row, max_col=max_column)
            cats = Reference(ws, min_col=rank_col, min_row=2, max_row=max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            # Modify each series in the chart to disable the inversion of negative values 
            for series in chart.series:
                series.invertIfNegative = False

            # y-axis ticks
            chart.y_axis.tickLblPos = "nextTo"
            chart.y_axis.delete = False  # Ensure axis is not deleted
            chart.y_axis.number_format = '0.000'

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Add the chart to the chart worksheet
            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)
            
            # Update position for the next chart
            current_col += chart_width +1
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)
        
    wb.save(filepath_workbook)

