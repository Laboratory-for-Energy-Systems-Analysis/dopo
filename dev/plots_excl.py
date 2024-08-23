# IN EXCEL
def _categorize_sheets_by_sector(file_path):
    # Load the workbook
    workbook = load_workbook(filename=file_path, read_only=True)
    
    # Initialize a dictionary to hold sectors and their corresponding sheet names
    worksheet_dict = {}
    
    # Iterate over all sheet names in the workbook
    for sheet_name in workbook.sheetnames:
        # Split the sheet name to extract the sector (assumes sector is the first part)
        sector = sheet_name.split('_')[0]
        
        # Add the sheet name to the corresponding sector in the dictionary
        if sector in worksheet_dict:
            worksheet_dict[sector].append(sheet_name)
        else:
            worksheet_dict[sector] = [sheet_name]
    
    return worksheet_dict


# ----
#PLOTS
# ----

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

def dot_plots(filepath_workbook, worksheet_dict, index_positions):

    worksheet_dict = _categorize_sheets_by_sector(filepath_workbook)
    
    # Load the workbook
    wb = load_workbook(filepath_workbook)
    
    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)        
                
        # Initial position for the first chart
        current_row = 1  # Start placing charts from row 1
        current_col = 1  # Start placing charts from column 1
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        charts_per_row = 3  # Number of charts per row
        
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find min_row, max_row and max_column
            max_row = ws.max_row
            max_column = ws.max_column
            min_row = 1

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in index_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = index_positions[matching_key]
            total_col = positions.get("total", None) + 1
            rank_col = positions.get("rank", None) + 1
            mean_col = positions.get("mean", None) + 1
            std_adv_col = positions.get("2std_abv", None) + 1
            std_blw_col = positions.get("2std_blw", None) + 1
            q1_col = positions.get("q1", None) + 1
            q3_col = positions.get("q3", None) + 1
            method_col = positions.get("method", None) + 1
            method_unit_col = positions.get("method unit", None) + 1
            
            # Ensure that all required columns are present
            if None in [total_col, rank_col, mean_col, std_adv_col, std_blw_col, q1_col, q3_col, method_col, method_unit_col]:
                print(f"Warning: Missing columns in worksheet '{worksheet_name}' for sector '{sector}'. Skipping...")
                continue
            
            # Create a ScatterChart (or other chart type as needed)
            chart = ScatterChart()

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"{method_value} LCA scores for {sector} sector" 
            
            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.y_axis.title = f"{method_unit_value}"
            chart.x_axis.title = 'activity rank'
            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 

            # Define the data range for the chart
            y_values = Reference(ws, min_col=total_col, min_row=min_row, max_row=max_row)
            x_values = Reference(ws, min_col=rank_col, min_row=min_row, max_row=max_row)

            # Create a series and add it to the chart
            series = Series(y_values, x_values, title_from_data=True)
            chart.series.append(series)
            chart.style = 9

            # Customize the series to show only markers (dots)
            series.marker.symbol = "circle"
            series.marker.size = 5
            series.graphicalProperties.line.noFill = True

            # ADJUST X-AXIS
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.majorGridlines = None 
            chart.x_axis.tickMarkSkip = 1  # Show all tick marks, this adresses the tick lines 
            chart.x_axis.tickLblSkip = 1  # Show all labels, doesnt work

            chart.x_axis.scaling.orientation = "minMax"
            chart.x_axis.crosses = "autoZero"
            chart.x_axis.axPos = "b"
            chart.x_axis.delete = False

            # ADJUST Y-AXIS
            chart.y_axis.tickLblPos = "nextTo"  # Position the labels next to the tick marks
            chart.y_axis.delete = False  # Ensure axis is not deleted
            chart.y_axis.number_format = '0.00000'
            chart.y_axis.majorGridlines = None 

            # ADD STATS
            # MEAN
            mean_y = Reference(ws, min_col=mean_col, min_row=min_row, max_row=max_row)
            mean_series = Series(mean_y, x_values, title_from_data="True")
            chart.series.append(mean_series)
            mean_series.marker.symbol = "none"  # No markers, just a line
            mean_series.graphicalProperties.line.solidFill = "FF0000"  # Red line for mean value
            mean_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)

            # IQR
            iqr1 = Reference(ws, min_col=q1_col, min_row=min_row, max_row=max_row)
            iqr3 = Reference(ws, min_col=q3_col, min_row=min_row, max_row=max_row)
            iqr1_series = Series(iqr1, x_values, title_from_data="True")
            iqr3_series = Series(iqr3, x_values, title_from_data="True")
            chart.series.append(iqr1_series)
            chart.series.append(iqr3_series)
            iqr1_series.marker.symbol = "none"  # No markers, just a line
            iqr3_series.marker.symbol = "none"
            iqr1_series.graphicalProperties.line.solidFill = "6082B6"  # Blue line 
            iqr3_series.graphicalProperties.line.solidFill = "6082B6"  
            iqr1_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)
            iqr3_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)

            # STD
            std_abv = Reference(ws, min_col=std_adv_col, min_row=min_row, max_row=max_row)
            std_blw = Reference(ws, min_col=std_blw_col, min_row=min_row, max_row=max_row)
            std_abv_series = Series(std_abv, x_values, title_from_data="True")
            std_blw_series = Series(std_blw, x_values, title_from_data="True")
            chart.series.append(std_abv_series)
            chart.series.append(std_blw_series)
            std_abv_series.marker.symbol = "none"  # No markers, just a line
            std_blw_series.marker.symbol = "none"
            std_abv_series.graphicalProperties.line.solidFill = "FFC300"  # yellow line
            std_blw_series.graphicalProperties.line.solidFill = "FFC300"  
            std_abv_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)
            std_blw_series.graphicalProperties.line.width = 10000  # Set line width (default units are EMUs)

            # Set legend position to the right of the plot area
            chart.legend.position = 'r'  # 'r' for right
            chart.legend.overlay = False

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)
            
            # Update position for the next chart
            current_col += chart_width +1 
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)

    wb.save(filepath_workbook)
    return current_row


from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def stacked_bars(filepath_workbook, worksheet_dict, index_positions, current_row_dot_plot):
    # Load the workbook
    wb = load_workbook(filepath_workbook)
    
    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)
                
        # Initial position for the first chart
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        current_row = current_row_dot_plot + chart_height # Start placing charts from row where dot plots have left of
        current_col = 1  # Start placing charts from column 1
        charts_per_row = 3  # Number of charts per row
        
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # Find the key in index_positions that contains worksheet_name
            matching_key = None
            for key in index_positions.keys():
                if worksheet_name in key:
                    matching_key = key
                    break

            if not matching_key:
                print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
                continue

            # Retrieve the column positions from the index_positions dictionary
            positions = index_positions[matching_key]

            # Find min_row, max_row and max_column
            max_row = ws.max_row
            max_column = ws.max_column
            input_min_col = positions.get("first_input", None) + 1
            rank_col = positions.get("rank", None) + 1
            method_col = positions.get("method", None) + 1
            method_unit_col = positions.get("method unit", None) + 1

            chart = BarChart()
            chart.type = "bar"
            chart.style = 2
            chart.grouping = "stacked"
            chart.overlap = 100

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"{sector} sector inputs contributions to {method_value}"

            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.y_axis.title = f"{method_unit_value}"
            
            chart.x_axis.title = 'activity index'

            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 
            chart.legend.overlay = False

            # Define data
            data = Reference(ws, min_col=input_min_col, min_row=1, max_row=max_row, max_col=max_column)
            cats = Reference(ws, min_col=rank_col, min_row=2, max_row=max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            # Modify each series in the chart to disable the inversion of negative values 
            for series in chart.series:
                series.invertIfNegative = False

            # y-axis ticks
            chart.y_axis.tickLblPos = "nextTo"
            chart.y_axis.delete = False  # Ensure axis is not deleted
            chart.y_axis.number_format = '0.000'

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Add the chart to the chart worksheet
            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)
            
            # Update position for the next chart
            current_col += chart_width +1
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)
        
    wb.save(filepath_workbook)

# Plot 3: Comparing databases

import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference

def compare_database_charts(filename, worksheet_dict, index_positions=None):

    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(filename)

    # Iterate over each sector and its associated worksheets
    for sector, worksheet_names in worksheet_dict.items():
        
        # Create or get the chart sheet for the current sector
        chart_sheet_name = f"{sector}_charts"
        if chart_sheet_name in wb.sheetnames:
            ws_charts = wb[chart_sheet_name]
        else:
            ws_charts = wb.create_sheet(chart_sheet_name)  
        
        # Initial position for the first chart
        current_row = 1  # Start placing charts from row 1
        current_col = 1  # Start placing charts from column 1
        chart_height = 30  # Number of rows a chart occupies
        chart_width = 12   # Number of columns a chart occupies
        charts_per_row = 2  # Number of charts per row
    
        # Iterate over each worksheet name in the current sector
        for i, worksheet_name in enumerate(worksheet_names):
            ws = wb[worksheet_name]

            # # Find the key in index_positions that contains worksheet_name
            # matching_key = None
            # for key in index_positions.keys():
            #     if worksheet_name in key:
            #         matching_key = key
            #         break

            # if not matching_key:
            #     print(f"Warning: No matching key found for worksheet '{worksheet_name}'. Skipping...")
            #     continue

            # Retrieve the column positions from the index_positions dictionary
            # positions = index_positions[matching_key]

            # Find min_row, max_row and max_column
            min_col_data = 15 #positions.get("relative_change", None) + 1
            rank_col = 17#positions.get("rank", None) + 1
            method_col = 5#positions.get("method", None) + 1
            method_unit_col = 6#positions.get("method unit", None) + 1

            # Create a bar chart
            chart = BarChart()
            chart.type="bar"
            chart.style=2
            chart.overlap= 100
            chart.title = "Relative Change in LCA Scores"
            chart.x_axis.title = "Activity"
            chart.y_axis.title = "Relative Change (%)"

            # Set the data for the chart
            data = Reference(ws, min_col=min_col_data, min_row=1, max_row=ws.max_row)
            categories = Reference(ws, min_col=rank_col, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # Modify each series in the chart to disable the inversion of negative values 
            for series in chart.series:
                series.invertIfNegative = False

            # x-axis tickes
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.majorGridlines = None 
            chart.x_axis.tickMarkSkip = 1  # Show all tick marks, this adresses the tick lines 
            chart.x_axis.tickLblSkip = 1  # Show all labels, doesnt work
            chart.x_axis.delete = False  # Ensure axis is not deleted

            # Chart titles
            method_value = ws.cell(row=2, column=method_col).value
            chart.title = f"{sector} {method_value} database lca scores relative changes"

            method_unit_value = ws.cell(row=2, column=method_unit_col).value
            chart.x_axis.title = f"{method_unit_value}"
            
            chart.y_axis.title = 'relative change (%)' #its switched..... should be x_axis

            # Avoid overlap
            chart.title.overlay = False
            chart.x_axis.title.overlay = False
            chart.y_axis.title.overlay = False 
            chart.legend.overlay = False

            # Adjust chart dimensions
            chart.width = 20  # Width of the chart
            chart.height = 14  # Height of the chart

            # Calculate the position for this chart
            position = ws_charts.cell(row=current_row, column=current_col).coordinate
            ws_charts.add_chart(chart, position)

            # Update position for the next chart
            current_col += chart_width +1 
            if (i + 1) % charts_per_row == 0:  # Move to the next row after placing `charts_per_row` charts
                current_row += chart_height +1
                current_col = 1  # Reset to the first column

        # Move the chart sheet to the first position
        wb._sheets.remove(ws_charts)
        wb._sheets.insert(0, ws_charts)

            # Add the chart to a new worksheet
            # new_sheet = wb.create_sheet(title="LCA Chart")
            # new_sheet.add_chart(chart, "A1")

    # Save the workbook
    wb.save(filename)

    print(f"Results and chart saved to {filename}")