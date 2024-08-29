import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series, BarChart

from dopo import (
    generate_sets_from_filters
)

import brightway2 as bw
import bw2data as bd
import bw2analyzer as ba

def sector_lca_scores(main_dict, method_dict, cutoff=0.02):
    '''
    Generates the LCA score tables for activity list of each sector.
    The tables contain total scores and cpc input contributions.
    This is done by each method defined in the method dictionary.

    :param main_dict: dictionary which is returned by process_yaml_files function
    :param method_dict: dictionary which is created with MethodFinder class
    :param cutoff: cutoff value to summarize inputs below or equal to this threshhold in a "other" column

    It returns the main dictionary updated as scores dictionary which also holds the former information for each sector.
    The LCA scores are stored by method name in the respective sector dictionary within the main dictionary.
    '''

    # Initialize scores_dict as a copy of main_dict
    scores_dict = main_dict.copy()

    # Loop through each sector in main_dict
    for sector in scores_dict.keys():
        # Extract activities for the current sector
        sector_activities = scores_dict[sector]['activities']
        
        # Calculate LCA scores using the specified method
        lca_scores = compare_activities_multiple_methods(
            activities_list=sector_activities,
            methods=method_dict,
            identifier=sector,
            mode='absolute'
        )
        
        # Apply the small_inputs_to_other_column function with the cutoff value
        lca_scores_cut = small_inputs_to_other_column(lca_scores, cutoff)
        
        # Save the LCA scores to the scores_dict
        scores_dict[sector]['lca_scores'] = lca_scores_cut

    return scores_dict


# Function based on brightways bw2analyzer (ba) function for generating dataframe containing total score and contribution by inputs
# -----------------------------------------------------------------------------------------------------------------------------


def compare_activities_multiple_methods(
    activities_list, methods, identifier, output_format="pandas", mode="absolute"
):
    """
    Compares a set of activities by multiple methods, stores each generated dataframe as a variable (the method is the variable name) in a dictionary.

    :param activities_list: List of activities to compare
    :param methods: List of Brightway Method objects
    :param identifier: A string used in defining the variable names to better identify comparisons (e.g. sector name).
    :param output_format: Output format for the comparison (default: 'pandas')
    :param mode: Mode for the comparison (default: 'absolute'; others: 'relative')
    :return: Dictionary of resulting dataframes from the comparisons
    """
    dataframes_dict = {}

    for method_key, method_details in methods.items():
        result = ba.comparisons.compare_activities_by_grouped_leaves(
            activities_list,
            method_details["object"].name,
            output_format=output_format,
            mode=mode,
        )

        # Create a variable name using the method name tuple and identifier
        method_name = method_details["object"].name[2].replace(" ", "_").lower()
        var_name = f"{identifier}_{method_name}"

        # add two columns method and method unit to the df
        result["method"] = str(method_details["object"].name[2])
        result["method unit"] = str(method_details["object"].metadata["unit"])

        # order the columns after column unit
        cols = list(result.columns)
        unit_index = cols.index("unit")
        cols.insert(unit_index + 1, cols.pop(cols.index("method")))
        cols.insert(unit_index + 2, cols.pop(cols.index("method unit")))
        result = result[cols]

        # Order the rows based on 'activity' and 'location' columns
        #result = result.sort_values(["activity", "location"])
        result=result.sort_values(['total'])

        # Reset the index numbering
        result = result.reset_index(drop=True)

        # Store the result in the dictionary
        dataframes_dict[var_name] = result

    return dataframes_dict


# Function for creating 'other' category for insignificant input contributions (for dataframes generated with compare_activities_multiple_methods)
# -------------------------------------------------------------------------------------------------------------------------------------------------

def small_inputs_to_other_column(dataframes_dict, cutoff=0.01):
    '''
    Aggregate values into a new 'other' column for those contributing less than or equal to the cutoff value to the 'total' column value.
    Set the aggregated values to zero in their original columns.
    Remove any columns that end up containing only zeros.

    Additionally, if a column is named None or "Unnamed", its values will be added to the 'other' column and then the original column will be deleted.

    :param dataframes_dict: the dictionary
    '''
    
    processed_dict = {}

    for key, df in dataframes_dict.items():
        # Identify the 'total' column
        total_col_index = df.columns.get_loc('total')
        
        # Separate string and numeric columns
        string_cols = df.iloc[:, :total_col_index]
        numeric_cols = df.iloc[:, total_col_index:]
        numeric_cols = numeric_cols.astype(float)
        
        # Create 'other' column
        numeric_cols['other'] = 0.0
        
        # Identify and handle columns that are None or called "Unnamed"
        columns_to_remove = []
        for col in df.columns:
            if col is None or col == "None" or str(col).startswith("Unnamed"):
                numeric_cols['other'] += df[col].fillna(0) 
                columns_to_remove.append(col)
        
        # Drop the identified columns
        numeric_cols.drop(columns=columns_to_remove, inplace=True)

        for col in numeric_cols.columns[1:-1]:  # Skip 'total' and 'other'
            mask_positive_total = numeric_cols['total'] > 0
            mask_negative_total = ~mask_positive_total
            
            # For rows with positive total values
            mask_pos = mask_positive_total & ((numeric_cols[col] < numeric_cols['total'] * cutoff) & 
                                            (numeric_cols[col] > numeric_cols['total'] * (-cutoff)))
            
            # For rows with negative total values
            mask_neg = mask_negative_total & ((numeric_cols[col] > numeric_cols['total'] * cutoff) & 
                                            (numeric_cols[col] < numeric_cols['total'] * (-cutoff)))
            
            # Apply the logic for both positive and negative totals
            numeric_cols.loc[mask_pos | mask_neg, 'other'] += numeric_cols.loc[mask_pos | mask_neg, col]
            numeric_cols.loc[mask_pos | mask_neg, col] = 0

            # Add these values to 'other'
            numeric_cols.loc[mask_pos, 'other'] += numeric_cols.loc[mask_pos, col]
            numeric_cols.loc[mask_neg, 'other'] += numeric_cols.loc[mask_neg, col]

            # Set these values to zero in the original column
            numeric_cols.loc[mask_pos, col] = 0
            numeric_cols.loc[mask_neg,col] = 0
        
        # Remove columns with all zeros (except 'total' and 'other')
        cols_to_keep = ['total'] + [col for col in numeric_cols.columns[1:-1] 
                                             if not (numeric_cols[col] == 0).all()]
        cols_to_keep.append('other')
        
        numeric_cols = numeric_cols[cols_to_keep]
        
        # Combine string and processed numeric columns
        processed_df = pd.concat([string_cols, numeric_cols], axis=1)
        
        # Sort DataFrame by total (optional)
        processed_df = processed_df.sort_values('total', ascending=False)
        
        # Store the processed DataFrame in the result dictionary
        processed_dict[key] = processed_df
        
    return processed_dict



#################################################################
#################################################################


import re
from openpyxl import load_workbook

def _add_statistics(df, column_name='total'):

    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions

    It adds statistical indicators to a dataframe based on total column which are used for plotting.

    returns updated dataframe
    '''
    
    #Need a rank row to plot the total LCA scores in descending order (satter opepyxl function takes in non categorial values)
    df['rank'] = df[column_name].rank(method="first", ascending=False)

    # Calculate mean, standard deviation, and IQR
    df['mean'] = df[column_name].mean()
    df['2std_abv'] = df['mean'] + df[column_name].std() * 2
    df['2std_blw'] = df['mean'] - df[column_name].std() * 2
    df['q1'] = df[column_name].quantile(0.25)
    df['q3'] = df[column_name].quantile(0.75)
    
    # Reorder the columns to place the new columns after 'total'
    cols = df.columns.tolist()
    total_index = cols.index(column_name) + 1
    new_cols = ['rank', 'mean', '2std_abv', '2std_blw', 'q1', 'q3']
    cols = cols[:total_index] + new_cols + cols[total_index:-len(new_cols)]
    
    return df[cols]


def _find_first_input_column(df):
    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions. Needs to be called before _clean_column_labels function.
    Detects the first column in the dataframe which contains input contribution data and saves its index. 
    This is relevant for calling the right column for defining the to be plotted data dynamically as not all dataframes have the same column order (some contain "direct emissions" for instance).
    '''
    
    def _clean_label(label):
        return label if label is not None else 'Unnamed'
    
    # Apply the cleaning function to all column names
    df.columns = [_clean_label(col) for col in df.columns]
    
    # Regular expression pattern to match "Number: Name"
    pattern = r'^\d+:\s*'
    
    for idx, column in enumerate(df.columns):
        if (column is not None and re.match(pattern, column)) or column == 'Unnamed' or column == 'direct emissions':
            return idx

    return None

def _clean_column_labels(df):

    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions. Needs to be run after _find_first_input_column.

    It removes unnecessary numbers in the column header.

    Returns df with formated column labels.
    '''
    # Function to remove numbers and colon from column names
    def _clean_label(label):
        if label is None:
            return 'Unnamed'  # or return 'Unnamed' if you prefer a placeholder
        return re.sub(r'^\d+:\s*', '', str(label))

    # Apply the cleaning function to all column names
    df.columns = [_clean_label(col) for col in df.columns]

    return df

def _add_sector_marker(df, sector):
    '''
    It is called in the function sector_lca_scores_to_excel_and_column_positions.

    It adds information about the sector for titel and labeling in plotting.

    Returns df with added column.
    '''
    
    # Add sector marker column
    df['sector']=str(sector) # potentially remove!
    # Reorder the columns to move 'sector' after 'product'
    columns = list(df.columns)

    if 'product' in df.columns:
        product_index = columns.index('product')
        # Insert 'sector' after 'product'
        columns.insert(product_index + 1, columns.pop(columns.index('sector')))
    else:
        # If 'product' does not exist, 'sector' remains in the last column
        columns.append(columns.pop(columns.index('sector')))
        
    # Reassign the DataFrame with the new column order
    df = df[columns]
    return df

def _categorize_sheets_by_sector(file_path):
    # Load the workbook
    workbook = load_workbook(filename=file_path, read_only=True)
    
    # Initialize a dictionary to hold sectors and their corresponding sheet names
    worksheet_dict = {}
    
    # Iterate over all sheet names in the workbook
    for sheet_name in workbook.sheetnames:
        # Skip combined sector sheets (assuming these sheets don't have an underscore)
        if '_' not in sheet_name:
            continue
        
        # Split the sheet name to extract the sector (assumes sector is the first part)
        sector = sheet_name.split('_')[0]
        
        # Add the sheet name to the corresponding sector in the dictionary
        if sector in worksheet_dict:
            worksheet_dict[sector].append(sheet_name)
        else:
            worksheet_dict[sector] = [sheet_name]
    
    return worksheet_dict


######################################
####################################

import pandas as pd

def sector_lca_scores_to_excel(scores_dict, excel_file_name):
    """ 
    What it does:
        - Creates a dataframe for each method and sector from the lca scores dictionary
        - Before storing each df in a worksheet in an excel file it:
                - shortens the column labels of the input (removing cpc code)
                - adds a sector name marker for keeping track in excel (when plotting can use it for labeling)
                - adds statistics for plotting
                - creates a dictionary which holds the indexes to the columns we need to call for plotting, this makes it dynamic. Otherwise need to hardcode index column number for openpxyl.
    What it returns:
        - Returns the index positions dictionary where the key is "sector_method"
        - Creates excel file as defined by user
    """

    # Dictionary to store positions of columns for each method
    column_positions = {}

    # DataFrames to store combined sector data
    combined_sector_dfs = {}
    method_dfs = []

    # Process each sector and its methods
    for sector in scores_dict.keys():
        sector_dfs = []
        lca_scores = scores_dict[sector]['lca_scores']

        # Process each method for the current sector
        for method, table in lca_scores.items():
            df = pd.DataFrame(table)

            # Add sector marker
            df = _add_sector_marker(df, sector)

            # Add statistics to the DataFrame
            df = _add_statistics(df)

            # Get the index values of columns
            columns_of_interest = ["total", "rank", "mean", "2std_abv", "2std_blw", "q1", "q3", "method", "method unit"]
            positions = {col: df.columns.get_loc(col) for col in columns_of_interest if col in df.columns}
            column_positions[f"{sector}_{method}"] = positions
            #column_positions[sector][method]=positions


            # Find the first input column and add it to the positions dictionary
            first_input_col_index = _find_first_input_column(df)
            if first_input_col_index is not None:
                positions["first_input"] = first_input_col_index

            # Store the positions for this method
            column_positions[f"{sector}_{method}"]= positions

            # Remove CPC from input labels
            df = _clean_column_labels(df)

            sector_dfs.append(df)

            # Store method-specific DataFrames for later
            method_dfs.append((f"{sector}_{method}", df))

        # Combine all dataframes for this sector
        combined_df = pd.concat(sector_dfs, axis=0, ignore_index=True, sort=False).fillna(0)
        combined_sector_dfs[sector] = combined_df

    # Write to Excel file
    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        # First write all combined sector sheets
        for sector, combined_df in combined_sector_dfs.items():
            worksheet_name_big = f"{sector}"
            if len(worksheet_name_big) > 31:
                worksheet_name_big = worksheet_name_big[:31]
            combined_df.to_excel(writer, sheet_name=worksheet_name_big, index=False)

        # Then write all method-specific sheets
        for worksheet_name, df in method_dfs:
            if len(worksheet_name) > 31:
                worksheet_name = worksheet_name[:31]
            df.to_excel(writer, sheet_name=worksheet_name, index=False)

    return column_positions

